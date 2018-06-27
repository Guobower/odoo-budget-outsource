# -*- coding: utf-8 -*-
import base64
import io

from odoo import models, fields, api
from odoo.exceptions import ValidationError

import pandas as pd
from openpyxl.styles import Protection

from .utils import *
from .utils import load_workbook
from .utils import odoo_to_pandas_list

from .sheet_mixin import MOBILIZE_COLUMNS


class DataSheet(models.Model):
    _name = 'budget.outsource.datasheet'
    _description = 'Data Sheet'
    _inherit = ['budget.outsource.sheet.mixin']

    # CHOICES
    # ----------------------------------------------------------

    # BASIC FIELDS
    # ----------------------------------------------------------

    # RELATIONSHIPS
    # ----------------------------------------------------------
    attendance_ids = fields.One2many('budget.outsource.mobilize.attendance',
                                     'datasheet_id',
                                     string='Attendances')
    sbh_ids = fields.One2many('budget.outsource.sbh',
                              'datasheet_id',
                              string='SBH')

    # MISC
    # ----------------------------------------------------------
    def get_querysets(self):
        querysets = super(DataSheet, self).get_querysets()
        q = [('is_mobilized', '=', True),
             ('resource_id.contractor_id', '=', self.contractor_id.id)]
        qs_mobilize = self.env['budget.outsource.mobilize'].search(q)
        querysets['qs_mobilize'] = qs_mobilize

        return querysets

    def create_attendances(self, context):
        qs_mobilize = context['qs_mobilize']
        mobilize_ids = qs_mobilize.mapped('id')
        attendance_ids = self.env['budget.outsource.mobilize.attendance'].search([
            ('period_start', '=', self.period_start),
            ('period_end', '=', self.period_end),
            ('mobilize_id', 'in', mobilize_ids),
            ('datasheet_id', '=', self.id),
        ])

        new_attendance_ids = set(mobilize_ids) - set(attendance_ids.mapped('mobilize_id.id'))
        # period_start, period_end, required_hours, mobilize_id, ,datasheet_id
        for i in new_attendance_ids:
            self.env['budget.outsource.mobilize.attendance'].create({
                'period_start': self.period_start,
                'period_end': self.period_end,
                'required_hours': context['required_hours'],
                'mobilize_id': i,
                'datasheet_id': self.id
            })

    def get_context(self):
        querysets = self.get_querysets()
        context = super(DataSheet, self).get_context()
        context['datasheet_template_path'] = os.path.join(context['form_template_path'], 'DATASHEET.xlsx')
        qs_mobilize = context['qs_mobilize'] = querysets['qs_mobilize']

        # DICTIONARY OF COLUMNS/FIELDS
        # (field_name, renamed_column)

        df_mobilize = pd.DataFrame(odoo_to_pandas_list(qs_mobilize, MOBILIZE_COLUMNS.keys()))
        # RENAME COLUMN NAMES TO FOR EASY USAGE
        df_mobilize.columns = [MOBILIZE_COLUMNS[i] for i in df_mobilize.columns]

        df_mobilize.sort_values(
            by=['division_alias', 'manager', 'director', 'position_name'],
            ascending=[True, True, True, True],
            inplace=True)

        rs_mobilize = df_mobilize.to_dict('records')
        context['rs_mobilize'] = rs_mobilize

        return context

    # BUTTONS
    # ----------------------------------------------------------
    @api.one
    def generate_datasheet(self):
        context = self.get_context()
        wb = load_workbook(context['datasheet_template_path'])

        # WORK SHEET - STAFF LIST
        # ---------------------------------------------------------------------------------------
        row = 7
        column = 2
        sr = 1
        ws = wb.get_sheet_by_name('STAFF LIST')

        ws.cell('B1').value = "713H CLAIMS FOR {}".format(context['contractor_name'])
        ws.cell('B2').value = context['period_string']

        rs_mobilize = context['rs_mobilize']

        ws.insert_rows(row, len(rs_mobilize) - 1)
        for record in rs_mobilize:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = record.get('division_alias', '')
            ws.cell(row=row, column=column + 2).value = record.get('section_name', '')
            ws.cell(row=row, column=column + 3).value = record.get('manager', '')
            ws.cell(row=row, column=column + 4).value = record.get('director', '')
            ws.cell(row=row, column=column + 5).value = record.get('resID', '')
            ws.cell(row=row, column=column + 6).value = record.get('positionID', '')
            ws.cell(row=row, column=column + 7).value = record.get('po_num', '')
            ws.cell(row=row, column=column + 8).value = record.get('unit_rate', 0.0)
            ws.cell(row=row, column=column + 9).value = record.get('rate_variance_percent', 0.0)
            ws.cell(row=row, column=column + 10).value = record.get('rate', 0.0)
            ws.cell(row=row, column=column + 11).value = record.get('agency_ref_num', '')
            ws.cell(row=row, column=column + 12).value = record.get('resource_name', '').title()
            ws.cell(row=row, column=column + 13).value = record.get('position_name', '')
            ws.cell(row=row, column=column + 14).value = record.get('position_level', '')
            ws.cell(row=row, column=column + 15).value = '' if not record['date_of_join'] else '{0:%d-%b-%Y}'.format(
                fields.Date.from_string(record['date_of_join']))
            ws.cell(row=row, column=column + 16).value = 'Yes' if record.get('has_tool_or_uniform', '') else ""
            ws.cell(row=row, column=column + 17).value = context['required_hours']
            ws.cell(row=row, column=column + 18).value = context['required_days']

            # # Unlock Cells
            ws.cell(row=row, column=column + 20).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 21).protection = Protection(locked=False)

            sr += 1
            row += 1

        row_auto_filter = row - 1
        ws.auto_filter.ref = "B6:T{}".format(row_auto_filter)
        ws.protection.autoFilter = False
        # ---------------------------------------------------------------------------------------
        # END WORK SHEET - STAFF LIST

        # WORK SHEET - MISMATCH
        # ---------------------------------------------------------------------------------------
        ws1 = wb.get_sheet_by_name('MISMATCH')
        ws1.auto_filter.ref = "B6:R{}".format(row_auto_filter)
        ws1.protection.autoFilter = False
        ws1.protection.insertRows = False
        # ---------------------------------------------------------------------------------------
        # END WORK SHEET - MISMATCH

        # APPLY PROTECTION IN THE WORKBOOK IN ALL SHEETS
        # ---------------------------------------------------------------------------------------
        for sheet in wb.worksheets:
            sheet.protection.enable()
            sheet.protection.set_password(context['xlsx_pass'])
        # ---------------------------------------------------------------------------------------
        # END

        tmp_file = io.BytesIO()
        wb.save(tmp_file)
        filename = 'DATASHEET ' + context['contractor_name'] + ' ' + context['period_string'] + '.xlsx'
        self.attach(filename, tmp_file)
        tmp_file.close()

        self.state = 'in progress'
        self.create_attendances(context)

    @api.one
    def analyze_datasheet(self):
        wb = load_workbook(filename=io.BytesIO(base64.b64decode(self.sheet_filled)), data_only=True)

        # WORK SHEET - STAFF LIST
        # ---------------------------------------------------------------------------------------
        row = 7
        mobilize_id_column = 'G'  # G
        rendered_hours_column = 'U'  # V
        remarks_column = 'W'  # W
        ws = wb.get_sheet_by_name('STAFF LIST')

        sr = 0
        for i in range(ws.min_row, ws.max_row):
            counter = ws.cell('B%s' % i).value
            sr += 0 if not isinstance(counter, int) else 1

        if sr != len(self.attendance_ids):
            raise ValidationError(
                'Count of filled Mobilization must be equal to the Mobilization Generated in the System')

        for i in range(0, sr):
            resource_identifier = ws.cell(mobilize_id_column + '{}'.format(row + i)).value
            rendered_hours = ws.cell(rendered_hours_column + '{}'.format(row + i)).value
            remarks = ws.cell(remarks_column + '{}'.format(row + i)).value

            attendance = self.attendance_ids.search([
                ('period_start', '=', self.period_start),
                ('period_end', '=', self.period_end),
                ('mobilize_id.resource_id.identifier', '=', int(resource_identifier)),
                ('datasheet_id', '=', self.id),
            ])

            attendance.write({
                'rendered_hours': rendered_hours,
                'remarks': remarks
            })

        self.state = 'approved'

    @api.one
    def generate_sbh(self):
        distincts = []
        mobilize_ids = self.attendance_ids.mapped('mobilize_id')
        for i in mobilize_ids:
            division_id = i.division_id.mapped('id')
            contractor_id = i.resource_id.mapped('contractor_id.id')
            po_id = i.position_id.mapped('po_id.id')

            query_params = division_id + contractor_id + po_id
            if len(query_params) < 3:
                continue

            record = f';'.join([str(j) for j in query_params])
            if record not in distincts:
                distincts += [record]

        for i in distincts:
            record = i.split(';')
            sbh_id = self.env['budget.outsource.sbh'].create({
                'division_id': record[0],
                'contractor_id': record[1],
                'po_id': record[2],
                'datasheet_id': self.id,
                'period_start': self.period_start,
                'period_end': self.period_end,
            })
            sbh_id.generate_sbh()

    @api.one
    def reset(self):
        self.attendance_ids.unlink()
        super(DataSheet, self).reset()

    @api.one
    def set_closed(self):
        states = self.attendance_ids.mapped('state')
        if 'approved' in states and len(states) == 1:
            self.state = 'closed'
        else:
            divisions = self.attendance_ids.filtered(lambda r: r.state != 'approved').mapped('division_id.name')
            ValidationError(f"All attendance should be approved, please check sbh for {', '.join(divisions)}")

    # ACTION METHODS
    # ----------------------------------------------------------
    @api.multi
    def action_show_attendance(self):
        domain = [('attendance_ids.datasheet_id', '=', self.id)]
        return super(DataSheet, self).action_show_attendance(domain)

    # POLYMORPH FUNCTIONS
    # ----------------------------------------------------------
    @api.multi
    def name_get(self):
        result = super(DataSheet, self).name_get()
        return [(i[0], 'DATASHEET ' + i[1]) for i in result]
