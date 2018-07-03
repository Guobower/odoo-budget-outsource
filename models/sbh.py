# -*- coding: utf-8 -*-
import base64
import io

from odoo import models, fields, api
from odoo.exceptions import ValidationError

import pandas as pd
from openpyxl.styles import Protection

from .utils import *
from .utils import load_workbook
from .utils import odoo_to_pandas_list

from .sheet_mixin import MOBILIZE_COLUMNS


class Sbh(models.Model):
    _name = 'budget.outsource.sbh'
    _description = 'SBH'
    _inherit = ['budget.outsource.sheet.mixin']

    # CHOICES
    # ----------------------------------------------------------

    # BASIC FIELDS
    # ----------------------------------------------------------

    # RELATIONSHIPS
    # ----------------------------------------------------------
    attendance_ids = fields.One2many('budget.outsource.mobilize.attendance',
                                     'sbh_id',
                                     string='Attendances')
    datasheet_id = fields.Many2one('budget.outsource.datasheet', string='DataSheet')

    # MISC
    # ----------------------------------------------------------
    def attach_attendance(self, attendance_ids):
        attendance_ids.write({
            'sbh_id': self.id
        })

    def get_querysets(self):
        querysets = super(Sbh, self).get_querysets()
        q = [('is_mobilized', '=', True),
             ('position_id.po_id', '=', self.po_id.id),
             ('resource_id.contractor_id', '=', self.contractor_id.id),
             ('position_id.division_id', '=', self.division_id.id)]

        qs_mobilize = self.env['budget.outsource.mobilize'].search(q)

        q = [('period_start', '=', self.period_start),
             ('period_end', '=', self.period_end),
             ('mobilize_id', 'in', qs_mobilize.ids)]
        qs_mobilize_attendance = self.env['budget.outsource.mobilize.attendance'].search(q)

        q = [('contractor_id', '=', self.contractor_id.id)]
        qs_unit_rate = self.env['budget.outsource.unit.rate'].search(q)

        querysets['qs_mobilize'] = qs_mobilize
        querysets['qs_mobilize_attendance'] = qs_mobilize_attendance
        querysets['qs_unit_rate'] = qs_unit_rate

        return querysets

    def get_context(self):
        querysets = self.get_querysets()
        context = super(Sbh, self).get_context()
        context['po_number'] = self.po_id.no
        context['sbh_template_path'] = os.path.join(context['form_template_path'], 'SBH-FORM.xlsx')

        qs_mobilize = context['qs_mobilize'] = querysets['qs_mobilize']
        qs_mobilize_attendance = context['qs_mobilize_attendance'] = querysets['qs_mobilize_attendance']
        qs_unit_rate = context['qs_unit_rate'] = querysets['qs_unit_rate']

        df_mobilize = pd.DataFrame(odoo_to_pandas_list(qs_mobilize, MOBILIZE_COLUMNS.keys()))
        # RENAME COLUMN NAMES TO FOR EASY USAGE
        df_mobilize.columns = [MOBILIZE_COLUMNS[i] for i in df_mobilize.columns]
        df_mobilize_attendance = pd.DataFrame(odoo_to_pandas_list(qs_mobilize_attendance, ['rendered_hours', 'remarks',
                                                                                           'mobilize_id.id']))
        df_mobilize = pd.merge(left=df_mobilize, right=df_mobilize_attendance,
                               how='left', left_on='id', right_on='mobilize_id.id')
        df_mobilize['period_start'] = context['period_start']
        df_mobilize['period_end'] = context['period_end']
        df_mobilize['required_hours'] = context['required_hours']

        df_mobilize.sort_values(
            by=['director', 'manager', 'position_name', 'rate_variance_percent'],
            ascending=[True, True, True, True],
            inplace=True)
        df_mobilize = df_mobilize.fillna(0.0)
        rs_mobilize = df_mobilize.to_dict('records')

        df_unit_rate = pd.DataFrame(odoo_to_pandas_list(qs_unit_rate))
        df_unit_rate['position_name'] = df_unit_rate['position_name'].apply(lambda x: x.upper())
        df_unit_rate['position_level'] = df_unit_rate['position_level'].apply(lambda x: x.lower())
        df_unit_rate = pd.pivot_table(df_unit_rate,
                                      index=['position_name'],
                                      columns='position_level',
                                      values="amount").reset_index()
        df_unit_rate = df_unit_rate.fillna(0.0)
        rs_unit_rate = df_unit_rate.to_dict('records')

        df_summary = pd.DataFrame(odoo_to_pandas_list(qs_mobilize, MOBILIZE_COLUMNS.keys()))
        # RENAME COLUMN NAMES TO FOR EASY USAGE
        df_summary.columns = [MOBILIZE_COLUMNS[i] for i in df_summary.columns]
        df_summary['position_name'] = df_summary['position_name'].apply(lambda x: x.upper())
        df_summary['position_level'] = df_summary['position_level'].apply(lambda x: x.lower())
        df_summary = df_summary.groupby(
            ['os_ref', 'position_name', 'position_level', 'rate_variance_percent']). \
            size().reset_index(name='count')
        df_summary = df_summary.pivot_table(
            index=['position_name', 'os_ref', 'rate_variance_percent'],
            columns='position_level',
            values='count')
        df_summary = df_summary.reset_index()
        df_summary.sort_values(by=['position_name', 'os_ref', 'rate_variance_percent'],
                               ascending=[True, True, True],
                               inplace=True)
        df_summary = df_summary.fillna(0.0)
        rs_summary = df_summary.to_dict('records')

        context['rs_mobilize'] = rs_mobilize
        context['rs_unit_price'] = rs_unit_rate
        context['rs_summary'] = rs_summary
        return context

    @api.one
    def generate_sbh(self):
        context = self.get_context()
        # Attach Attendance
        # ---------------------------------------------------------------------------------------
        self.attach_attendance(context['qs_mobilize_attendance'])

        wb = load_workbook(context['sbh_template_path'])

        # WORK SHEET SBH-FORM 1
        # ---------------------------------------------------------------------------------------
        row = 15
        column = 2
        sr = 1
        ws = wb.get_sheet_by_name('SBH-FORM 1')

        ws.cell('D5').value = context['contractor_name']
        ws.cell('G5').value = context['po_number']
        ws.cell('G7').value = 1
        ws.cell('D7').value = context['period_string']

        # CELLS TO BE UNLOCK
        unlock_cells = [
            'J39', 'L39',
            'J41', 'L41',
            'J43',

            'G46', 'J46', 'L46', 'M46',

            'C49', 'D49',

            'C53', 'C53',
            'C54', 'C54',
            'C55', 'C55',
            'C56', 'C56',
        ]

        rs_mobilize = context['rs_mobilize']

        ws.insert_rows(row, len(rs_mobilize) - 1)
        for record in rs_mobilize:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = record['os_ref']
            ws.cell(row=row, column=column + 2).value = record['agency_ref_num']
            ws.cell(row=row, column=column + 3).value = record['resource_name']
            ws.cell(row=row, column=column + 4).value = record['position_name']
            ws.cell(row=row, column=column + 5).value = record['position_level']
            ws.cell(row=row, column=column + 6).value = '' if not record['date_of_join'] else '{0:%d-%b-%Y}'.format(
                fields.Date.from_string(record['date_of_join']))
            ws.cell(row=row, column=column + 7).value = record['rate_variance_percent']
            ws.cell(row=row, column=column + 8).value = record['required_hours']
            ws.cell(row=row, column=column + 9).value = record['rendered_hours']
            ws.cell(row=row, column=column + 10).value = 'Yes' if record['has_tool_or_uniform'] else ''
            ws.cell(row=row, column=column + 11).value = record['remarks'] or ''
            ws.cell(row=row, column=column + 12).value = record['manager'] or ''
            ws.cell(row=row, column=column + 13).value = record['director'] or ''
            ws.cell(row=row, column=column + 14).value = ''

            # # Unlock Cells
            ws.cell(row=row, column=column + 9).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 11).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 12).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 13).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 14).protection = Protection(locked=False)

            sr += 1
            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 1

        # WORK SHEET SBH-FORM 2
        # ---------------------------------------------------------------------------------------
        row = 16
        column = 2  # B

        ws = wb.get_sheet_by_name('SBH-FORM 2')
        rs_summary = context['rs_summary']
        ws.insert_rows(row, len(rs_summary) - 1)
        ws.cell('R13').value = round(context['required_hours'])

        for record in rs_summary:
            ws.cell(row=row, column=column).value = record['os_ref']
            ws.cell(row=row, column=column + 1).value = record.get('position_name')
            ws.cell(row=row, column=column + 2).value = record.get('level 1', 0)
            ws.cell(row=row, column=column + 3).value = record.get('level 2', 0)
            ws.cell(row=row, column=column + 4).value = record.get('level 3', 0)
            ws.cell(row=row, column=column + 12).value = record.get('rate_variance_percent', 0)

            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 2

        # WORK SHEET UNIT PRICE
        # ---------------------------------------------------------------------------------------
        row = 1
        column = 1  # B

        ws = wb.get_sheet_by_name('UNIT PRICE')
        rs_unit_price = context['rs_unit_price']
        for record in rs_unit_price:
            ws.cell(row=row, column=column).value = record['position_name']
            ws.cell(row=row, column=column + 1).value = record.get('level 1', 0)
            ws.cell(row=row, column=column + 2).value = record.get('level 2', 0)
            ws.cell(row=row, column=column + 3).value = record.get('level 3', 0)
            ws.cell(row=row, column=column + 4).value = record.get('level 4', 0)

            row += 1
        # ---------------------------------------------------------------------------------------
        # END UNIT PRICE

        # UNLOCK CELLS
        # ---------------------------------------------------------------------------------------
        for unlock_cell in unlock_cells:
            u_cell = ws.cell(unlock_cell)
            ws.cell(row=u_cell.row + sr, column=u_cell.col_idx).protection = Protection(locked=False)

        # ---------------------------------------------------------------------------------------
        # END UNLOCK CELLS

        # APPLY PROTECTION IN THE WORKBOOK IN ALL SHEETS
        # ---------------------------------------------------------------------------------------
        for sheet in wb.worksheets:
            sheet.protection.enable()
            sheet.protection.set_password(context['xlsx_pass'])
        # ---------------------------------------------------------------------------------------
        # END

        tmp_file = io.BytesIO()
        wb.save(tmp_file)

        filename = '{:%m} {} {} {}.xlsx'.format(context['period_start'],
                                                context['contractor_name'],
                                                context['po_number'],
                                                context['period_string'])
        self.attach(filename, tmp_file)
        tmp_file.close()

        self.state = 'in progress'

    @api.one
    def analyse_sbh(self):
        pass

    # BUTTONS
    # ----------------------------------------------------------
    @api.one
    def set_approved(self):
        self.attendance_ids.write({'is_approved': True})
        self.state = 'approved'

    @api.multi
    def download_sbh(self):
        attachment_id = self.env['ir.attachment'].search([
            ('res_model', '=', self._name),
            ('res_id', '=', self.id)
        ], limit=1, order='id desc')

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment_id.id,
            'target': 'self',
            'res_id': self.id,
        }

    # ACTION METHODS
    # ----------------------------------------------------------
    @api.multi
    def action_show_attendance(self):
        domain = [('attendance_ids.sbh_id', '=', self.id)]
        return super(Sbh, self).action_show_attendance(domain)
